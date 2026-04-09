"""Excel-Dateien (.xlsx, .xls, .csv) einlesen."""

from pathlib import Path
import openpyxl
import csv


def parse_excel(file_path: Path) -> list[dict]:
    """Excel-Datei einlesen. Gibt Tabellenblaetter zurueck."""
    suffix = file_path.suffix.lower()

    if suffix == ".csv":
        return _parse_csv(file_path)
    return _parse_xlsx(file_path)


def _parse_xlsx(file_path: Path) -> list[dict]:
    """XLSX/XLS einlesen."""
    wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
    results = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            row_data = [str(cell) if cell is not None else "" for cell in row]
            if any(cell.strip() for cell in row_data):
                rows.append(row_data)

        if rows:
            # Als Text formatieren
            text_lines = [" | ".join(row) for row in rows]
            results.append({
                "text": "\n".join(text_lines),
                "content_type": "table",
                "section": sheet_name,
                "table_data": rows,
                "row_count": len(rows),
                "col_count": len(rows[0]) if rows else 0,
            })

    wb.close()
    return results


def _parse_csv(file_path: Path) -> list[dict]:
    """CSV einlesen."""
    rows = []
    # Encoding erraten
    for encoding in ["utf-8", "latin-1", "cp1252"]:
        try:
            with open(file_path, "r", encoding=encoding) as f:
                reader = csv.reader(f)
                rows = [row for row in reader if any(cell.strip() for cell in row)]
            break
        except UnicodeDecodeError:
            continue

    if not rows:
        return []

    text_lines = [" | ".join(row) for row in rows]
    return [{
        "text": "\n".join(text_lines),
        "content_type": "table",
        "section": file_path.stem,
        "table_data": rows,
        "row_count": len(rows),
    }]
